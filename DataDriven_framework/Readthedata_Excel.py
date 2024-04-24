import openpyxl

wb = openpyxl.load_workbook("C:\\Users\\Harshavardhan\\OneDrive\\Desktop\\DDF1.xlsx")

sh = wb['Sheet1']

print("total rows in excel sheet1 = ",sh.max_row)

print("total columns in excel sheet1 = ", sh.max_column)

print(sh.cell(row=6,column=2).value)



print("============")
for r in range(1,sh.max_row+1):

    for c in range(1, sh.max_column+1):

        print(sh.cell(row=r,column=c).value)
