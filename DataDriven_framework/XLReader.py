import openpyxl


def total_rows(excelPath, sheetName):
    wb = openpyxl.load_workbook(excelPath)
    sh = wb[sheetName]
    rows = sh.max_row
    return rows


def total_columns(excelPath, sheetName):
    wb = openpyxl.load_workbook(excelPath)
    sh = wb[sheetName]
    cols = sh.max_column
    return cols


def read_data(excelPath, sheetName, r, c):
    wb = openpyxl.load_workbook(excelPath)
    sh = wb[sheetName]
    data = sh.cell(row=r,column=c).value
    return data


def write_data(excelPath, sheetName, r, c, val):
    wb = openpyxl.load_workbook(excelPath)
    sh = wb[sheetName]
    sh.cell(row=r,column=c).value=val
    wb.save(excelPath)
